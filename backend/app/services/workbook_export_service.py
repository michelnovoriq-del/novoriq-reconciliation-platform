import io
import json
import uuid
from datetime import datetime, timezone
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import ClientWorkspace, Organization, ReconciliationRun, RejectedRecord, User
from app.services.matching_service import _load_results
from app.utils.money import parse_money

HEADER_FILL = PatternFill("solid", fgColor="123B5D")
HEADER_FONT = Font(color="FFFFFF", bold=True)

def _value(value):
    if value is None: return ""
    if isinstance(value, (dict, list)): return json.dumps(value, sort_keys=True, default=str)
    if isinstance(value, uuid.UUID): return str(value)
    if isinstance(value, datetime) and value.tzinfo: return value.astimezone(timezone.utc).replace(tzinfo=None)
    return value

def _table(sheet, headers: list[str], rows: Iterable[list], *, empty_note: str | None = None) -> None:
    sheet.append(headers)
    for cell in sheet[1]: cell.fill = HEADER_FILL; cell.font = HEADER_FONT
    sheet.freeze_panes = "A2"; sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    row_count = 0
    for row in rows:
        sheet.append([_value(value) for value in row]); row_count += 1
    if not row_count and empty_note:
        sheet.append([empty_note] + [""] * (len(headers) - 1))
    for index, header in enumerate(headers, 1): sheet.column_dimensions[get_column_letter(index)].width = min(max(len(header) + 2, 14), 34)
    for column_index, header in enumerate(headers, 1):
        normalized = header.casefold()
        for cell in sheet.iter_cols(min_col=column_index, max_col=column_index, min_row=2):
            for value_cell in cell:
                if "date" in normalized or normalized.endswith("_at"): value_cell.number_format = "yyyy-mm-dd"
                elif "amount" in normalized or "difference" in normalized or "fee" in normalized: value_cell.number_format = '#,##0.00;[Red]-#,##0.00'
                elif header == "confidence_score": value_cell.number_format = '0"%"'

def _exception_type(status: str) -> str:
    return {
        "possible_match": "unclear_reference", "amount_variance": "amount_variance",
        "late_settlement": "late_settlement", "duplicate_candidate": "duplicate_candidate",
        "manual_review_required": "manual_review_required", "unmatched_file_a": "unmatched_payment",
        "unmatched_file_b": "unmatched_bank_deposit", "rejected": "manual_review_required",
    }.get(status, status)

def _recommended(status: str) -> str:
    return {
        "amount_variance": "Review processor fees, adjustments, refunds, or rounding differences.", "late_settlement": "Confirm whether this payout settled outside the normal date tolerance.",
        "duplicate_candidate": "Multiple candidate matches exist. Review manually before approval.", "possible_match": "Review source rows because the reference is missing or unclear.",
        "manual_review_required": "Review the supporting source evidence.", "rejected": "Review the rejection reason and source rows.",
        "unmatched_file_a": "Check whether the bank statement is missing the related deposit or whether the payout settled later.",
        "unmatched_file_b": "Check whether this deposit belongs to another processor, transfer, refund, or manual credit.",
    }.get(status, "Review the source evidence before resolving this exception.")

def _format_duration(minutes: int) -> str:
    if minutes < 60: return f"{minutes} minute{'s' if minutes != 1 else ''}"
    hours, remainder = divmod(minutes, 60)
    return f"{hours} hour{'s' if hours != 1 else ''}" + (f" {remainder} minutes" if remainder else "")

def build_workbook_filename(db: Session, *, run: ReconciliationRun, workspace_name: str | None = None) -> str:
    results = _load_results(db, run.id, run.organization_id)
    dates = [record.transaction_date for item in results for record in (item.file_a_record, item.file_b_record) if record and record.transaction_date]
    period = min(dates).strftime("%Y-%m") if dates else datetime.now(timezone.utc).strftime("%Y-%m-%d")
    workspace = workspace_name or ""
    safe_workspace = "_".join(filter(None, ("".join(character if character.isalnum() else " " for character in workspace)).split()))
    parts = ["Novoriq_Reconciliation_Report"]
    if safe_workspace: parts.append(safe_workspace)
    parts.extend((period, str(run.id)[:8]))
    return "_".join(parts) + ".xlsx"

def build_reconciliation_workbook(db: Session, *, run: ReconciliationRun, organization: Organization) -> bytes:
    results = _load_results(db, run.id, organization.id)
    workspace = db.get(ClientWorkspace, run.workspace_id) if run.workspace_id else None
    creator = db.get(User, run.created_by_user_id)
    rejected = list(db.scalars(select(RejectedRecord).where(RejectedRecord.organization_id == organization.id, RejectedRecord.uploaded_file_id.in_([run.file_a_id, run.file_b_id]))))
    suggested_counts = {status: sum(item.suggested_status == status for item in results) for status in {item.suggested_status for item in results}}
    suggested_confident = sum(item.suggested_status == "confident_match" for item in results)
    approved_count = sum(item.status == "approved" for item in results)
    pending_confident = sum(item.suggested_status == "confident_match" and item.status != "approved" for item in results)
    pending_review = sum(not item.reviewed_at for item in results)
    exception_count = sum(item.suggested_status in {"possible_match","amount_variance","late_settlement","duplicate_candidate","manual_review_required","currency_mismatch","unclear_reference"} for item in results)
    workbook = Workbook(); summary = workbook.active; summary.title = "Summary"
    summary["A1"] = "Novoriq Reconciliation Summary"; summary["A1"].font = Font(size=18, bold=True, color="123B5D")
    summary_rows = [
        ("Client / Workspace", workspace.name if workspace else "Unassigned"), ("Organization", organization.name), ("Run ID", str(run.id)),
        ("Created At", run.created_at), ("Created By", creator.email if creator else str(run.created_by_user_id)), ("Review Status", "reviewed" if any(x.reviewed_at for x in results) else "pending_review"),
        ("Payment File", run.file_a.original_filename), ("Bank File", run.file_b.original_filename),
        ("Payment Mapping Used", run.file_a.normalization_mapping or {}), ("Bank Mapping Used", run.file_b.normalization_mapping or {}),
        ("Total Payment Rows", run.file_a.row_count or 0), ("Total Bank Rows", run.file_b.row_count or 0), ("Total Rows Processed", (run.file_a.row_count or 0) + (run.file_b.row_count or 0)),
        ("Suggested Confident Matches", suggested_confident), ("Approved Matches", approved_count), ("Pending Confident Matches", pending_confident), ("Pending Review", pending_review), ("Exceptions", exception_count), ("Possible Matches", sum(item.suggested_status == "possible_match" for item in results)),
        ("Amount Variances", suggested_counts.get("amount_variance", 0)), ("Late Settlements", suggested_counts.get("late_settlement", 0)), ("Duplicate Candidates", suggested_counts.get("duplicate_candidate", 0)),
        ("Unmatched Payment Rows", suggested_counts.get("unmatched_file_a", 0)), ("Unmatched Bank Rows", suggested_counts.get("unmatched_file_b", 0)), ("Invalid Rows", len(rejected)),
        ("Manual Review Required", suggested_counts.get("manual_review_required", 0)), ("Estimated Time Saved", _format_duration(round((run.file_a.row_count or 0) / 100 * 20))),
        ("Export Generated At", datetime.now(timezone.utc)),
    ]
    for row_index, (label, value) in enumerate(summary_rows, 3): summary.cell(row_index, 1, label).font = Font(bold=True); summary.cell(row_index, 2, _value(value))
    summary.cell(len(summary_rows) + 5, 1, "Suggested confident matches are deterministic matches identified by Novoriq. Approved matches are matches reviewed and approved by a user.")
    summary.column_dimensions["A"].width = 30; summary.column_dimensions["B"].width = 80

    match_headers = ["match_id","status","suggested_status","review_status","confidence_score","match_reason","recommended_action","payment_row_number","payment_date","payment_reference","payment_order_id","payment_customer_name","payment_processor","payment_gross_amount","payment_processor_fee","payment_net_amount","payment_amount_used","payment_currency","payment_status","bank_row_number","bank_date","bank_reference","bank_description","bank_amount","bank_currency","bank_account_number","bank_transaction_type","amount_difference","date_difference_days","approved_by","approved_at","reviewed_by","reviewed_at","review_notes"]
    def match_row(item):
        a=item.file_a_record; b=item.file_b_record; ar=a.raw_data if a else {}; br=b.raw_data if b else {}
        review_status="approved" if item.status=="approved" else "rejected" if item.status=="rejected" else "pending_review"
        return [item.id,item.status,item.suggested_status,review_status,item.confidence_score,item.match_reason,"Approve after confirming the source evidence.",a.source_row_number if a else None,a.transaction_date if a else None,a.reference if a else None,ar.get("order_id"),a.customer_name if a else None,ar.get("processor"),parse_money(ar.get("gross_amount")),parse_money(ar.get("processor_fee")),parse_money(ar.get("net_amount")),a.amount if a else None,a.currency if a else None,ar.get("status"),b.source_row_number if b else None,b.transaction_date if b else None,b.reference if b else None,b.description if b else None,b.amount if b else None,b.currency if b else None,br.get("account_number"),br.get("transaction_type"),item.amount_difference,item.date_difference_days,item.reviewed_by_user_id if item.status=="approved" else None,item.reviewed_at if item.status=="approved" else None,item.reviewed_by_user_id,item.reviewed_at,item.review_notes]
    suggested = workbook.create_sheet("Suggested Confident Matches"); _table(suggested, match_headers, (match_row(x) for x in results if x.suggested_status == "confident_match"))
    approved = workbook.create_sheet("Approved Matches"); _table(approved, match_headers, (match_row(x) for x in results if x.status == "approved"), empty_note="No approved matches yet. Suggested confident matches are available in the Suggested Confident Matches sheet.")

    exceptions = workbook.create_sheet("Exceptions"); exception_headers=["exception_id","exception_type","risk_level","status","review_status","confidence_score","match_reason","recommended_action","payment_row_number","payment_date","payment_reference","payment_order_id","payment_amount","payment_currency","payment_status","bank_row_number","bank_date","bank_reference","bank_description","bank_amount","bank_currency","amount_difference","date_difference_days","review_notes","reviewed_by","reviewed_at"]
    exception_statuses={"possible_match","amount_variance","late_settlement","duplicate_candidate","manual_review_required","currency_mismatch","unclear_reference","rejected"}
    def exception_row(x):
        a=x.file_a_record;b=x.file_b_record;ar=a.raw_data if a else {}
        return [x.id,_exception_type(x.suggested_status),"high" if x.suggested_status in {"duplicate_candidate","currency_mismatch"} else "medium",x.status,"reviewed" if x.reviewed_at else "pending_review",x.confidence_score,x.match_reason,_recommended(x.suggested_status),a.source_row_number if a else None,a.transaction_date if a else None,a.reference if a else None,ar.get("order_id"),a.amount if a else None,a.currency if a else None,ar.get("status"),b.source_row_number if b else None,b.transaction_date if b else None,b.reference if b else None,b.description if b else None,b.amount if b else None,b.currency if b else None,x.amount_difference,x.date_difference_days,x.review_notes,x.reviewed_by_user_id,x.reviewed_at]
    _table(exceptions, exception_headers, (exception_row(x) for x in results if x.suggested_status in exception_statuses))

    unmatched_headers=["record_id","row_number","date","reference","description","amount_used","currency","unmatched_reason","recommended_action","review_status","reviewed_by","reviewed_at","review_notes"]
    for title,status,side in (("Unmatched Payment Rows","unmatched_file_a","file_a_record"),("Unmatched Bank Rows","unmatched_file_b","file_b_record")):
        sheet=workbook.create_sheet(title)
        def rows(status=status,side=side):
            for x in results:
                if x.suggested_status != status: continue
                record=getattr(x,side); yield [record.id,record.source_row_number,record.transaction_date,record.reference,record.description,record.amount,record.currency,x.match_reason,_recommended(status),"reviewed_unmatched" if x.status=="reviewed_unmatched" else "pending_review",x.reviewed_by_user_id,x.reviewed_at,x.review_notes]
        _table(sheet,unmatched_headers,rows())

    mapping=workbook.create_sheet("Mapping Audit"); mapping_headers=["file_id","file_name","file_type","uploaded_at","mapped_date_column","mapped_amount_column","mapped_reference_column","mapped_description_column","mapped_currency_column","mapped_customer_column","normalized_row_count","mapping_source","mapping_json"]
    _table(mapping,mapping_headers,([f.id,f.original_filename,f.file_type,f.created_at,(f.normalization_mapping or {}).get("date"),(f.normalization_mapping or {}).get("amount"),(f.normalization_mapping or {}).get("reference"),(f.normalization_mapping or {}).get("description"),(f.normalization_mapping or {}).get("currency"),(f.normalization_mapping or {}).get("customer_name"),f.row_count,"manual_or_suggested",f.normalization_mapping or {}] for f in (run.file_a,run.file_b)))
    invalid=workbook.create_sheet("Invalid Rows"); invalid_headers=["file_id","file_name","row_number","field_name","raw_value","error_reason","recommended_fix","created_at"]
    file_names={run.file_a_id:run.file_a.original_filename,run.file_b_id:run.file_b.original_filename}; invalid_rows=[]
    for item in rejected:
        errors=item.field_errors or {"row":item.rejection_reason}
        for field,reason in errors.items(): invalid_rows.append([item.uploaded_file_id,file_names.get(item.uploaded_file_id,""),item.source_row_number,field,(item.raw_data or {}).get(field),reason,"Correct the source value and normalize the file again.",item.created_at])
    _table(invalid,invalid_headers,invalid_rows or [["","","","","","No invalid rows found.","",""]])
    output=io.BytesIO(); workbook.save(output); return output.getvalue()
