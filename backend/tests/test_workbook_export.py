import io
import uuid
from datetime import date, datetime, timezone
from decimal import Decimal
from unittest.mock import MagicMock

from openpyxl import load_workbook

from app.models import ClientWorkspace, MatchResult, NormalizedRecord, Organization, ReconciliationRun, UploadedFile, User
from app.services.workbook_export_service import _exception_type, _recommended, build_reconciliation_workbook, build_workbook_filename
from app.routes import reconciliation_runs as reconciliation_routes

SHEETS = ["Summary","Suggested Confident Matches","Approved Matches","Exceptions","Unmatched Payment Rows","Unmatched Bank Rows","Mapping Audit","Invalid Rows"]

def fixture_objects(status: str = "confident_match"):
    organization=Organization(id=uuid.uuid4(),name="Demo Firm"); user=User(id=uuid.uuid4(),email="owner@example.com",hashed_password="test")
    workspace=ClientWorkspace(id=uuid.uuid4(),organization_id=organization.id,name="Demo Ecommerce Client",slug="demo-ecommerce-client",status="active",created_by_user_id=user.id)
    payment=UploadedFile(id=uuid.uuid4(),organization_id=organization.id,workspace_id=workspace.id,uploaded_by_user_id=user.id,original_filename="novoriq_payment_export_1000_rows.csv",stored_filename="payment.csv",file_path="payment.csv",file_type="csv",status="normalized",row_count=1000,normalization_mapping={"date":"payout_date","amount":"net_amount","reference":"payout_reference"})
    bank=UploadedFile(id=uuid.uuid4(),organization_id=organization.id,workspace_id=workspace.id,uploaded_by_user_id=user.id,original_filename="novoriq_bank_statement_1000_rows.csv",stored_filename="bank.csv",file_path="bank.csv",file_type="csv",status="normalized",row_count=1000,normalization_mapping={"date":"value_date","amount":"deposit_amount","reference":"bank_reference"})
    run=ReconciliationRun(id=uuid.uuid4(),organization_id=organization.id,workspace_id=workspace.id,created_by_user_id=user.id,file_a_id=payment.id,file_b_id=bank.id,status="completed",created_at=datetime(2026,7,18,tzinfo=timezone.utc));run.file_a=payment;run.file_b=bank
    left=NormalizedRecord(id=uuid.uuid4(),source_row_number=1,transaction_date=date(2026,7,2),amount=Decimal("695.96"),reference="PAY-1",currency="USD",raw_data={"net_amount":"695.96","gross_amount":"709.64","processor_fee":"13.68"})
    right=NormalizedRecord(id=uuid.uuid4(),source_row_number=1,transaction_date=date(2026,7,4),amount=Decimal("695.96"),reference="PAY-1",currency="USD",raw_data={"account_number":"ACC-1"})
    result=MatchResult(id=uuid.uuid4(),reconciliation_run_id=run.id,organization_id=organization.id,file_a_record_id=left.id,file_b_record_id=right.id,status=status,suggested_status="confident_match",confidence_score=100,match_reason="Exact net amount and reference.",amount_difference=Decimal("0"),date_difference_days=2,reviewed_by_user_id=user.id if status=="approved" else None,reviewed_at=datetime.now(timezone.utc) if status=="approved" else None);result.file_a_record=left;result.file_b_record=right
    db=MagicMock();db.get.side_effect=lambda model,value:workspace if model is ClientWorkspace else user;db.scalars.return_value=[]
    return db,organization,workspace,run,result

def summary_values(workbook):
    return {workbook["Summary"].cell(row,1).value:workbook["Summary"].cell(row,2).value for row in range(3,workbook["Summary"].max_row+1)}

def test_pre_review_workbook_shows_suggested_matches_and_empty_approved_note(monkeypatch):
    db,organization,workspace,run,result=fixture_objects()
    monkeypatch.setattr("app.services.workbook_export_service._load_results",lambda *args,**kwargs:[result])
    workbook=load_workbook(io.BytesIO(build_reconciliation_workbook(db,run=run,organization=organization)))
    assert workbook.sheetnames==SHEETS
    assert workbook["Suggested Confident Matches"].max_row==2
    assert workbook["Approved Matches"].max_row==2
    assert "No approved matches yet" in workbook["Approved Matches"][2][0].value
    summary=summary_values(workbook)
    assert summary["Suggested Confident Matches"]==1
    assert summary["Approved Matches"]==0
    assert summary["Pending Confident Matches"]==1
    assert summary["Estimated Time Saved"]=="3 hours 20 minutes"
    headers=[cell.value for cell in workbook["Suggested Confident Matches"][1]]
    row=[cell.value for cell in workbook["Suggested Confident Matches"][2]]
    assert row[headers.index("payment_amount_used")]==row[headers.index("payment_net_amount")]==695.96

def test_approved_match_remains_suggested_and_appears_in_approved_sheet(monkeypatch):
    db,organization,workspace,run,result=fixture_objects("approved")
    monkeypatch.setattr("app.services.workbook_export_service._load_results",lambda *args,**kwargs:[result])
    workbook=load_workbook(io.BytesIO(build_reconciliation_workbook(db,run=run,organization=organization)))
    assert workbook["Suggested Confident Matches"].max_row==2
    assert workbook["Approved Matches"].max_row==2
    assert workbook["Approved Matches"][2][0].value==str(result.id)
    assert summary_values(workbook)["Approved Matches"]==1

def test_mapping_audit_and_filename_are_accountant_friendly(monkeypatch):
    db,organization,workspace,run,result=fixture_objects()
    monkeypatch.setattr("app.services.workbook_export_service._load_results",lambda *args,**kwargs:[result])
    workbook=load_workbook(io.BytesIO(build_reconciliation_workbook(db,run=run,organization=organization)))
    rows=list(workbook["Mapping Audit"].iter_rows(values_only=True));amount_column=rows[0].index("mapped_amount_column")
    assert [row[amount_column] for row in rows[1:]]==["net_amount","deposit_amount"]
    assert workbook["Invalid Rows"][2][5].value=="No invalid rows found."
    assert build_workbook_filename(db,run=run,workspace_name=workspace.name)==f"Novoriq_Reconciliation_Report_Demo_Ecommerce_Client_2026-07_{str(run.id)[:8]}.xlsx"

def test_review_categories_have_explicit_exception_types_and_actions():
    assert _exception_type("possible_match")=="unclear_reference"
    assert _exception_type("unmatched_file_a")=="unmatched_payment"
    assert _exception_type("unmatched_file_b")=="unmatched_bank_deposit"
    for status in ("possible_match","amount_variance","late_settlement","duplicate_candidate","unmatched_file_a","unmatched_file_b"):
        assert _recommended(status)

def test_live_export_route_returns_corrected_workbook_and_filename(monkeypatch):
    db,organization,workspace,run,result=fixture_objects()
    run.workspace=workspace
    monkeypatch.setattr("app.services.workbook_export_service._load_results",lambda *args,**kwargs:[result])
    monkeypatch.setattr(reconciliation_routes,"get_reconciliation_run",lambda *args,**kwargs:run)

    response=reconciliation_routes.export_run_workbook(run.id,db=db,current_user=User(id=run.created_by_user_id,email="owner@example.com",hashed_password="x"),organization=organization)
    workbook=load_workbook(io.BytesIO(response.body))

    assert workbook.sheetnames==SHEETS
    assert workbook["Suggested Confident Matches"].max_row==2
    assert response.headers["x-novoriq-workbook-layout"]=="2"
    assert f"Demo_Ecommerce_Client_2026-07_{str(run.id)[:8]}.xlsx" in response.headers["content-disposition"]
